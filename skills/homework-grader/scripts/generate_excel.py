"""
Generate final graded Excel from JSON grading data.
Applies normal distribution curve and formatting.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel(grades_json, output_xlsx):
    """Generate Excel from JSON grading data."""
    with open(grades_json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    N = len(data)
    data.sort(key=lambda x: sum(x['scores']), reverse=True)
    
    # Normal distribution curve
    t5 = round(N * 0.15)
    g11 = round(N * 0.35)
    m11 = round(N * 0.35)
    
    def curve(rank, n):
        if rank < t5: return round(93 - (rank/max(1,t5-1))*3)
        elif rank < t5+g11: return round(89 - ((rank-t5)/max(1,g11-1))*9)
        elif rank < t5+g11+m11: return round(79 - ((rank-t5-g11)/max(1,m11-1))*9)
        else: return round(69 - ((rank-t5-g11-m11)/max(1,n-t5-g11-m11-1))*9)
    
    for i, s in enumerate(data):
        c = curve(i, N)
        s['curved'] = c
        s['grade'] = '优秀' if c>=90 else '良好' if c>=80 else '中等' if c>=70 else '及格'
    
    # Sort by student ID for Excel
    data.sort(key=lambda x: int(x['id']))
    
    wb = Workbook()
    ws = wb.active
    ws.title = '成绩表'
    
    headers = ['学号','姓名','总分','等级'] + \
              ['项目功能实现与运行效果(10分)','知识整合应用(10分)','编码及报告规范(10分)',
               '逻辑严谨性(10分)','思维链工具使用(10分)','迭代优化痕迹(10分)',
               '人机协同调试能力(10分)','演示表达与成果展示(10分)',
               '批判性甄别与合规意识(10分)','心得体会与素养反思(10分)'] + \
              ['优点','缺点及改进建议','指导教师评语']
    
    hfont = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cfont = Font(name='Arial', size=9)
    calign = Alignment(vertical='top', wrap_text=True)
    ccenter = Alignment(horizontal='center', vertical='top')
    bdr = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    fills = {
        '优秀': PatternFill('solid',fgColor='C6EFCE'),
        '良好': PatternFill('solid',fgColor='BDD7EE'),
        '中等': PatternFill('solid',fgColor='FFF2CC'),
        '及格': PatternFill('solid',fgColor='F4B4C2'),
        '不及格': PatternFill('solid',fgColor='FF6B6B'),
    }
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bdr
    
    widths = [6,8,6,8] + [16]*10 + [35,40,25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    for ri, s in enumerate(data, 2):
        vals = [s['id'], s['name'], s['curved'], s['grade']] + s['scores'] + \
               [s.get('advantages',''), s.get('disadvantages',''), s.get('comment','')]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font, c.border = cfont, bdr
            c.alignment = ccenter if col <= 4 or (5 <= col <= 14) else calign
            if col <= 4:
                c.fill = fills.get(s['grade'], fills['不及格'])
    
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:Q{N+1}'
    for row in range(2, N+2):
        ws.row_dimensions[row].height = 90
    ws.row_dimensions[1].height = 40
    
    wb.save(output_xlsx)
    print(f'Saved: {output_xlsx}')
    print(f'Students: {N}')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python generate_excel.py <grades.json> <output.xlsx>")
        sys.exit(1)
    generate_excel(sys.argv[1], sys.argv[2])
